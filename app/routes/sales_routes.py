from datetime import datetime, date
import os
from io import BytesIO
import csv
import json
from itsdangerous import URLSafeSerializer, BadSignature
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, make_response, current_app, abort
from flask_login import login_required, current_user
from app import db
from app.whatsapp_service import send_whatsapp_text, send_whatsapp_template, build_quotation_message, normalize_whatsapp_number, extract_whatsapp_message_id, whatsapp_config
from app.models import Customer, Project, ProjectType, User, SalesInquiry, SalesQuotation, SalesQuotationLine, QuotationItem, NotificationLog, WhatsAppMessage, INQUIRY_STATUSES, INQUIRY_SOURCES, QUOTATION_STATUSES

sales_bp = Blueprint('sales', __name__, url_prefix='/sales')

def parse_date(v):
    if not v: return None
    return datetime.strptime(v, '%Y-%m-%d').date()

def parse_float(v):
    try: return float(v or 0)
    except Exception: return 0

def make_ref(prefix):
    return f"{prefix}-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

@sales_bp.before_app_request
def ensure_sales_tables():
    try: db.create_all()
    except Exception: pass

@sales_bp.route('/')
@login_required
def dashboard():
    inquiries=SalesInquiry.query.order_by(SalesInquiry.created_at.desc()).limit(10).all()
    quotations=SalesQuotation.query.order_by(SalesQuotation.created_at.desc()).limit(10).all()
    summary={
        'inquiries': SalesInquiry.query.count(),
        'new': SalesInquiry.query.filter_by(status='New').count(),
        'quoted': SalesInquiry.query.filter_by(status='Quoted').count(),
        'won': SalesInquiry.query.filter_by(status='Won').count(),
        'quotations': SalesQuotation.query.count(),
        'items': QuotationItem.query.filter_by(is_active=True).count(),
    }
    return render_template('sales/dashboard.html', inquiries=inquiries, quotations=quotations, summary=summary)

@sales_bp.route('/quotation-items', methods=['GET','POST'])
@login_required
def quotation_items():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        added = 0
        updated = 0
        if uploaded and uploaded.filename:
            filename = uploaded.filename.lower()
            rows = []
            if filename.endswith('.xlsx'):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(uploaded, data_only=True)
                    ws = wb.active
                    headers = [str(c.value or '').strip().lower().replace(' ', '_') for c in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        data = dict(zip(headers, row))
                        rows.append(data)
                except Exception as e:
                    flash(f'Excel import failed: {e}', 'danger')
                    return redirect(url_for('sales.quotation_items'))
            elif filename.endswith('.csv'):
                content = uploaded.stream.read().decode('utf-8-sig').splitlines()
                reader = csv.DictReader(content)
                rows = [{(k or '').strip().lower().replace(' ', '_'): v for k, v in r.items()} for r in reader]
            else:
                flash('Please upload .xlsx or .csv file.', 'warning')
                return redirect(url_for('sales.quotation_items'))
            for r in rows:
                item_name = str(r.get('item_name') or r.get('item') or r.get('name') or '').strip()
                if not item_name: continue
                item = QuotationItem.query.filter_by(item_name=item_name).first()
                if not item:
                    item = QuotationItem(item_name=item_name)
                    db.session.add(item); added += 1
                else:
                    updated += 1
                item.description = str(r.get('description') or r.get('desc') or item.description or '').strip()
                item.unit = str(r.get('unit') or item.unit or '').strip()
                item.unit_price = parse_float(r.get('unit_price') or r.get('price') or item.unit_price)
                item.category = str(r.get('category') or item.category or '').strip()
                item.is_active = True
            db.session.commit()
            flash(f'Quotation items imported. Added: {added}, Updated: {updated}', 'success')
        else:
            item_name=request.form.get('item_name','').strip()
            if not item_name:
                flash('Item name is required.', 'warning')
            else:
                item=QuotationItem.query.filter_by(item_name=item_name).first() or QuotationItem(item_name=item_name)
                item.description=request.form.get('description')
                item.unit=request.form.get('unit')
                item.unit_price=parse_float(request.form.get('unit_price'))
                item.category=request.form.get('category')
                item.is_active=True
                db.session.add(item); db.session.commit()
                flash('Quotation item saved.', 'success')
        return redirect(url_for('sales.quotation_items'))
    items=QuotationItem.query.order_by(QuotationItem.category.asc(), QuotationItem.item_name.asc()).all()
    return render_template('sales/quotation_items.html', items=items)

@sales_bp.route('/quotation-items/template')
@login_required
def quotation_items_template():
    try:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title='Quotation Items'
        ws.append(['item_name','description','unit','unit_price','category'])
        ws.append(['Solar Panel','High efficiency PV solar module','pcs',0,'PV Module'])
        ws.append(['Hybrid Inverter','Hybrid solar inverter','pcs',0,'Inverter'])
        bio=BytesIO(); wb.save(bio); bio.seek(0)
        return send_file(bio, as_attachment=True, download_name='quotation_items_template.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception:
        csv_data='item_name,description,unit,unit_price,category\nSolar Panel,High efficiency PV solar module,pcs,0,PV Module\n'
        return send_file(BytesIO(csv_data.encode()), as_attachment=True, download_name='quotation_items_template.csv', mimetype='text/csv')

@sales_bp.route('/inquiries')
@login_required
def inquiries():
    status=request.args.get('status')
    q=request.args.get('q','').strip()
    query=SalesInquiry.query
    if status: query=query.filter_by(status=status)
    if q:
        like=f'%{q}%'
        query=query.filter((SalesInquiry.customer_name.ilike(like)) | (SalesInquiry.ref_no.ilike(like)) | (SalesInquiry.phone.ilike(like)))
    inquiries=query.order_by(SalesInquiry.created_at.desc()).all()
    return render_template('sales/inquiries.html', inquiries=inquiries, statuses=INQUIRY_STATUSES, filters=request.args)

@sales_bp.route('/inquiries/create', methods=['GET','POST'])
@login_required
def create_inquiry():
    if request.method=='POST':
        customer_id=request.form.get('customer_id') or None
        customer=Customer.query.get(customer_id) if customer_id else None
        customer_name=request.form.get('customer_name') or (customer.customer_name if customer else '')
        if customer_name and not Customer.query.filter_by(customer_name=customer_name).first():
            db.session.add(Customer(customer_name=customer_name, phone=request.form.get('phone'), customer_type='Sales Lead', is_active=True))
        inquiry=SalesInquiry(ref_no=make_ref('INQ'), inquiry_date=parse_date(request.form.get('inquiry_date')) or date.today(), customer_id=customer_id, customer_name=customer_name, phone=request.form.get('phone'), location=request.form.get('location'), source=request.form.get('source'), project_type=request.form.get('project_type'), requirement_summary=request.form.get('requirement_summary'), estimated_capacity=request.form.get('estimated_capacity'), status=request.form.get('status') or 'New', assigned_to_id=request.form.get('assigned_to_id') or None, created_by_id=current_user.id, next_followup_date=parse_date(request.form.get('next_followup_date')), notes=request.form.get('notes'))
        db.session.add(inquiry); db.session.commit()
        flash(f'Inquiry {inquiry.ref_no} created.', 'success')
        return redirect(url_for('sales.inquiry_detail', inquiry_id=inquiry.id))
    return render_template('sales/inquiry_form.html', customers=Customer.query.order_by(Customer.customer_name.asc()).all(), project_types=ProjectType.query.order_by(ProjectType.type_name.asc()).all(), users=User.query.filter_by(is_active=True).all(), statuses=INQUIRY_STATUSES, sources=INQUIRY_SOURCES)

@sales_bp.route('/inquiries/<int:inquiry_id>')
@login_required
def inquiry_detail(inquiry_id):
    inquiry=SalesInquiry.query.get_or_404(inquiry_id)
    return render_template('sales/inquiry_detail.html', inquiry=inquiry)

@sales_bp.route('/quotations')
@login_required
def quotations():
    quotations=SalesQuotation.query.order_by(SalesQuotation.created_at.desc()).all()
    return render_template('sales/quotations.html', quotations=quotations)



def _customer_phone_for_quotation(quotation):
    """Best-effort lookup of the customer's WhatsApp/phone number for a quotation.

    The quotation table stores customer_name as text, so this lookup is made
    deliberately tolerant: inquiry phone first, then exact trimmed customer
    name, then case-insensitive customer name. This makes the WhatsApp number
    box auto-fill even when the customer was selected from the quotation
    dropdown but spacing/case differs slightly.
    """
    try:
        inquiry_phone = (getattr(getattr(quotation, 'inquiry', None), 'phone', '') or '').strip()
        if inquiry_phone:
            return inquiry_phone
    except Exception:
        pass

    customer_name = (getattr(quotation, 'customer_name', '') or '').strip()
    if not customer_name:
        return ''

    try:
        customer = Customer.query.filter(Customer.customer_name == customer_name).first()
        if not customer:
            customer = Customer.query.filter(Customer.customer_name.ilike(customer_name)).first()
        if customer:
            phone = (getattr(customer, 'phone', '') or '').strip()
            if phone:
                return phone
    except Exception:
        pass
    return ''


def _quotation_public_serializer():
    """Serializer used to create secure customer quotation links without login."""
    return URLSafeSerializer(current_app.config.get('SECRET_KEY', 'change-this-secret-key'), salt='quotation-public-link')


def _quotation_public_token(quotation):
    """Create a signed token tied to the quotation id and ref_no.

    The token is URL-safe and contains the quotation id internally, so the
    customer link can be short: /sales/q/<token>. This avoids Meta's dynamic URL
    limitation that only one variable may be added at the end of a button URL.
    """
    return _quotation_public_serializer().dumps({
        'qid': quotation.id,
        'ref': quotation.ref_no,
    })


def _quotation_public_token_value(quotation):
    """Return the signed token only, without a URL prefix."""
    return _quotation_public_token(quotation)


def _quotation_public_base_url():
    """Return the production-safe public base URL.

    This intentionally prefers PUBLIC_BASE_URL instead of request.host_url so
    locally generated WhatsApp links do not become http://127.0.0.1:5000/...
    """
    cfg = whatsapp_config()
    return (cfg.get('public_base_url') or os.environ.get('PUBLIC_BASE_URL') or 'https://cmse2.onrender.com').rstrip('/')


def _quotation_template_name():
    """Approved Meta template used for quotation messages.

    WHATSAPP_TEMPLATE_NAME remains available for the manual cmse_test page.
    Quotations use WHATSAPP_QUOTATION_TEMPLATE_NAME or quotation_ready_v2.
    """
    return (os.environ.get('WHATSAPP_QUOTATION_TEMPLATE_NAME') or 'quotation_ready_v2').strip()


def _quotation_token_data(token):
    """Decode a public quotation token and return its data or None."""
    try:
        return _quotation_public_serializer().loads(token or '')
    except BadSignature:
        return None


def _verify_quotation_public_token(quotation, token):
    """Validate a public quotation token. Returns True/False."""
    data = _quotation_token_data(token)
    if not data:
        return False
    return int(data.get('qid') or 0) == int(quotation.id) and data.get('ref') == quotation.ref_no


def _quotation_public_url(quotation):
    """Full public customer-facing quotation URL without requiring ERP login.

    Always uses PUBLIC_BASE_URL (default: https://cmse2.onrender.com) instead
    of request.host_url. This prevents WhatsApp links generated during local
    testing from becoming http://127.0.0.1:5000/...
    """
    token = _quotation_public_token_value(quotation)
    return f"{_quotation_public_base_url()}{url_for('sales.public_quotation_short', token=token)}"


def _quotation_form_context(**extra):
    context = dict(
        projects=Project.query.order_by(Project.project_name.asc()).all(),
        customers=Customer.query.filter_by(is_active=True).order_by(Customer.customer_name.asc()).all(),
        project_types=ProjectType.query.order_by(ProjectType.type_name.asc()).all(),
        statuses=QUOTATION_STATUSES,
        quotation_items=QuotationItem.query.filter_by(is_active=True).order_by(QuotationItem.item_name.asc()).all(),
    )
    context.update(extra)
    return context


def _save_quotation_lines(quotation):
    # Rebuild quotation lines from the editable form. This keeps edit simple and reliable.
    SalesQuotationLine.query.filter_by(quotation_id=quotation.id).delete()
    db.session.flush()
    for idx, item in enumerate(request.form.getlist('item'), start=1):
        if not item:
            continue
        db.session.add(SalesQuotationLine(
            quotation_id=quotation.id,
            item=item,
            description=request.form.get(f'description_{idx}'),
            quantity=parse_float(request.form.get(f'quantity_{idx}')),
            unit=request.form.get(f'unit_{idx}'),
            unit_price=parse_float(request.form.get(f'unit_price_{idx}')),
        ))

@sales_bp.route('/quotations/create', methods=['GET','POST'])
@login_required
def create_quotation():
    inquiry_id=request.args.get('inquiry_id') or request.form.get('inquiry_id') or None
    inquiry=SalesInquiry.query.get(inquiry_id) if inquiry_id else None
    if request.method=='POST':
        quotation=SalesQuotation(ref_no=make_ref('QTN'), inquiry_id=inquiry_id, project_id=request.form.get('project_id') or None, quotation_date=parse_date(request.form.get('quotation_date')) or date.today(), customer_name=request.form.get('customer_name'), project_type=request.form.get('project_type'), capacity=request.form.get('capacity'), scope_of_work=request.form.get('scope_of_work'), validity_days=int(request.form.get('validity_days') or 15), status=request.form.get('status') or 'Draft', prepared_by_id=current_user.id, notes=request.form.get('notes'))
        db.session.add(quotation); db.session.flush()
        _save_quotation_lines(quotation)
        if inquiry:
            inquiry.status='Quoted'
        db.session.commit(); flash(f'Quotation {quotation.ref_no} created.', 'success')
        return redirect(url_for('sales.edit_quotation', quotation_id=quotation.id))
    return render_template('sales/quotation_form.html', **_quotation_form_context(inquiry=inquiry, quotation=None, mode='create', max_rows=15))

@sales_bp.route('/quotations/<int:quotation_id>')
@login_required
def quotation_detail(quotation_id):
    quotation=SalesQuotation.query.get_or_404(quotation_id)
    customer_phone = _customer_phone_for_quotation(quotation)
    return render_template('sales/quotation_detail.html', quotation=quotation, customer_phone=customer_phone)

@sales_bp.route('/quotations/<int:quotation_id>/edit', methods=['GET','POST'])
@login_required
def edit_quotation(quotation_id):
    quotation=SalesQuotation.query.get_or_404(quotation_id)
    inquiry=quotation.inquiry
    if request.method=='POST':
        quotation.project_id=request.form.get('project_id') or None
        quotation.quotation_date=parse_date(request.form.get('quotation_date')) or quotation.quotation_date or date.today()
        quotation.customer_name=request.form.get('customer_name')
        quotation.project_type=request.form.get('project_type')
        quotation.capacity=request.form.get('capacity')
        quotation.scope_of_work=request.form.get('scope_of_work')
        quotation.validity_days=int(request.form.get('validity_days') or 15)
        quotation.status=request.form.get('status') or 'Draft'
        quotation.notes=request.form.get('notes')
        _save_quotation_lines(quotation)
        db.session.commit(); flash(f'Quotation {quotation.ref_no} updated.', 'success')
        return redirect(url_for('sales.edit_quotation', quotation_id=quotation.id))
    max_rows=max(15, len(quotation.lines)+5)
    return render_template('sales/quotation_form.html', **_quotation_form_context(inquiry=inquiry, quotation=quotation, mode='edit', max_rows=max_rows))



@sales_bp.route('/quotations/<int:quotation_id>/whatsapp', methods=['POST'])
@login_required
def send_quotation_whatsapp(quotation_id):
    quotation = SalesQuotation.query.get_or_404(quotation_id)
    phone = (request.form.get('phone') or '').strip() or _customer_phone_for_quotation(quotation)
    if not phone:
        flash('No phone number provided and no customer phone found. Please enter a WhatsApp number.', 'danger')
        return redirect(url_for('sales.quotation_detail', quotation_id=quotation.id))

    customer_name = (quotation.customer_name or 'Customer').strip()
    quotation_ref = (quotation.ref_no or f'Quotation-{quotation.id}').strip()
    quotation_token = _quotation_public_token_value(quotation)
    quotation_link = _quotation_public_url(quotation)

    # quotation_ready_v2 Meta template:
    # Body variables:
    #   {{1}} Customer Name
    #   {{2}} Quotation Number
    # Button:
    #   Dynamic URL configured in Meta as https://cmse2.onrender.com/sales/q/{{1}}
    #   ERP sends only the signed token as the button variable.
    template_name = _quotation_template_name()
    message = (
        f"Hello {customer_name}\n\n"
        f"Your quotation from Cadceed-Maal Solar Energy is ready.\n\n"
        f"Quotation Number: {quotation_ref}\n\n"
        f"Please click the Open Quotation button to view your quotation.\n\n"
        f"Public Link: {quotation_link}"
    )

    ok, response = send_whatsapp_template(
        phone,
        template_name=template_name,
        language_code=os.environ.get('WHATSAPP_QUOTATION_TEMPLATE_LANGUAGE', 'en'),
        components=[
            {
                'type': 'body',
                'parameters': [
                    {'type': 'text', 'text': customer_name},
                    {'type': 'text', 'text': quotation_ref},
                ],
            },
            {
                'type': 'button',
                'sub_type': 'url',
                'index': '0',
                'parameters': [
                    {'type': 'text', 'text': quotation_token},
                ],
            },
        ],
    )
    log = NotificationLog(
        recipient_user_id=None,
        recipient_name=quotation.customer_name,
        channel='WhatsApp',
        recipient=normalize_whatsapp_number(phone),
        subject=f'Quotation {quotation.ref_no}',
        message=message,
        related_module='Sales Quotation',
        related_ref=quotation.ref_no,
        status='Sent' if ok else 'Failed',
        provider_response=json.dumps(response)[:4000],
        error_message=None if ok else json.dumps(response)[:1000],
        sent_at=datetime.utcnow() if ok else None,
        created_by_id=current_user.id,
    )
    db.session.add(log)

    whatsapp_log = WhatsAppMessage(
        direction='Outbound',
        recipient_name=quotation.customer_name,
        phone_number=normalize_whatsapp_number(phone),
        template_name=template_name,
        message_body=message,
        variables_json=json.dumps({
            'customer_name': customer_name,
            'quotation_ref': quotation_ref,
            'quotation_token': quotation_token,
            'quotation_link': quotation_link,
        })[:3000],
        document_type='Sales Quotation',
        document_id=quotation.id,
        document_ref=quotation.ref_no,
        meta_message_id=extract_whatsapp_message_id(response),
        status='Sent' if ok else 'Failed',
        error_message=None if ok else json.dumps(response)[:1000],
        provider_response=json.dumps(response)[:4000],
        sent_by_id=current_user.id,
        sent_at=datetime.utcnow() if ok else None,
        failed_at=datetime.utcnow() if not ok else None,
    )
    db.session.add(whatsapp_log)
    db.session.commit()
    if ok:
        flash(f'Quotation {quotation.ref_no} sent to WhatsApp number {phone}.', 'success')
    else:
        flash(f'WhatsApp sending failed: {response}', 'danger')
    return redirect(url_for('sales.quotation_detail', quotation_id=quotation.id))



@sales_bp.route('/q/<token>')
def public_quotation_short(token):
    """Short public customer quotation view opened from WhatsApp without ERP login.

    This route opens the signed public quotation link without ERP login.
    The WhatsApp button receives the complete URL, for example:
        https://cmse2.onrender.com/sales/q/<signed-token>
    """
    data = _quotation_token_data(token)
    if not data:
        abort(404)
    quotation = SalesQuotation.query.get_or_404(int(data.get('qid') or 0))
    if data.get('ref') != quotation.ref_no:
        abort(404)
    return render_template('sales/public_quotation.html', quotation=quotation, generated_at=datetime.utcnow())


@sales_bp.route('/q/<token>/pdf')
def public_quotation_short_pdf(token):
    """Public PDF view/download for the signed customer quotation link."""
    data = _quotation_token_data(token)
    if not data:
        abort(404)
    quotation = SalesQuotation.query.get_or_404(int(data.get('qid') or 0))
    if data.get('ref') != quotation.ref_no:
        abort(404)
    html = render_template('sales/quotation_pdf.html', quotation=quotation, generated_at=datetime.utcnow())
    try:
        from weasyprint import HTML
        pdf = HTML(string=html, base_url=request.host_url).write_pdf()
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename={quotation.ref_no}.pdf'
        return response
    except Exception:
        return render_template('sales/public_quotation.html', quotation=quotation, generated_at=datetime.utcnow(), pdf_error=True)


@sales_bp.route('/public/quotations/<int:quotation_id>/<token>')
def public_quotation(quotation_id, token):
    """Public customer quotation view opened from WhatsApp without ERP login."""
    quotation = SalesQuotation.query.get_or_404(quotation_id)
    if not _verify_quotation_public_token(quotation, token):
        abort(404)
    return render_template('sales/public_quotation.html', quotation=quotation, generated_at=datetime.utcnow())

@sales_bp.route('/quotations/<int:quotation_id>/pdf')
@login_required
def quotation_pdf(quotation_id):
    quotation=SalesQuotation.query.get_or_404(quotation_id)
    html=render_template('sales/quotation_pdf.html', quotation=quotation, generated_at=datetime.utcnow())
    try:
        from weasyprint import HTML
        pdf=HTML(string=html, base_url=request.host_url).write_pdf()
        response=make_response(pdf)
        response.headers['Content-Type']='application/pdf'
        response.headers['Content-Disposition']=f'inline; filename={quotation.ref_no}.pdf'
        return response
    except Exception as e:
        flash(f'PDF generation failed: {e}', 'danger')
        return redirect(url_for('sales.quotation_detail', quotation_id=quotation.id))
